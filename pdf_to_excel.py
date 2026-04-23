"""
PDF Trade Report → Excel Portfolio Report converter
PASHA Kapital trade report → PASHA Private Banking portfolio format

Usage:
    python pdf_to_excel.py <input.pdf> <template.xlsx> [output.xlsx]
"""

import sys
import re
import os
import pdfplumber
import openpyxl


# ── Security definitions ────────────────────────────────────────────────────
# Each entry: (canonical_name, pdf_keywords, excel_keywords)
# pdf_keywords   — must appear in the PDF row text to identify the security
# excel_keywords — must appear in the Excel cell to match the row

SECURITIES = [
    (
        "SPDR S&P 500 ETF TRUST (SPY)",
        ["S&P", "SPY"],
        ["SPY", "S&P 500"],
    ),
    (
        "SPDR GOLD SHARES (GLD)",
        ["GOLD SHARES", "GLD"],
        ["GLD", "GOLD SHARES"],
    ),
    (
        "VANGUARD MID-CAP ETF (VO)",
        ["MID-CAP", "VANGUARD"],
        ["MID-CAP", "VO)"],
    ),
    (
        "iShares 7-10 YEAR TREASURY BOND (IEF)",
        ["7-10", "IEF"],
        ["IEF", "7-10"],
    ),
    (
        "iSHARES 1-5Y INVESTMENT GRADE CORP BOND",
        ["1-5Y", "INV GRADE", "GRADE CORP"],
        ["GRADE CORP", "1-5Y", "IGIB"],
    ),
]


def identify_security_from_text(text):
    """Return canonical name if text contains PDF keywords for a security."""
    text_upper = text.upper()
    for canonical, pdf_kws, _ in SECURITIES:
        for kw in pdf_kws:
            if kw.upper() in text_upper:
                return canonical
    return None


def excel_row_matches(cell_value, canonical_name):
    """Return True if the Excel cell matches the canonical security name."""
    if not cell_value:
        return False
    cell_upper = str(cell_value).upper()
    for name, _, excel_kws in SECURITIES:
        if name == canonical_name:
            for kw in excel_kws:
                if kw.upper() in cell_upper:
                    return True
    return False


# ── Number parsing ──────────────────────────────────────────────────────────

def parse_number(text):
    """Parse '(31,083.76)' or '31,083.76' or '31083.76' → float or None."""
    if not text:
        return None
    text = str(text).strip()
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "").replace(" ", "")
    try:
        val = float(text)
        return -val if negative else val
    except ValueError:
        return None


def group_words_into_rows(words, y_tolerance=4):
    """Group pdfplumber words into rows by similar Y coordinate."""
    if not words:
        return []
    rows = []
    current_row = [words[0]]
    current_y = words[0]["top"]
    for word in words[1:]:
        if abs(word["top"] - current_y) <= y_tolerance:
            current_row.append(word)
        else:
            rows.append(sorted(current_row, key=lambda w: w["x0"]))
            current_row = [word]
            current_y = word["top"]
    rows.append(sorted(current_row, key=lambda w: w["x0"]))
    return rows


# ── PDF extraction ──────────────────────────────────────────────────────────

def extract_pdf_data(pdf_path):
    """
    Extract holdings data from the PASHA Kapital trade report PDF.
    Uses word-coordinate approach instead of table parsing for reliability.
    """
    holdings = {}   # canonical_name → dict of values

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=15, y_tolerance=3,
                                       keep_blank_chars=False)
            if not words:
                continue

            words_sorted = sorted(words, key=lambda w: (round(w["top"] / 3) * 3, w["x0"]))
            rows = group_words_into_rows(words_sorted, y_tolerance=5)

            for i, row in enumerate(rows):
                row_text = " ".join(w["text"] for w in row)
                sec_name = identify_security_from_text(row_text)
                if not sec_name:
                    continue

                # Collect words from this row and the next 2 rows
                # (multi-line table cells sometimes split across rows)
                combined_words = list(row)
                for j in range(i + 1, min(i + 3, len(rows))):
                    next_text = " ".join(w["text"] for w in rows[j])
                    if identify_security_from_text(next_text):
                        break   # stop if next row is a different security
                    combined_words.extend(rows[j])

                # Extract all positive numbers with their X positions
                # (negative amounts appear in brackets which we handle)
                num_positions = []
                for w in combined_words:
                    n = parse_number(w["text"])
                    if n is not None:
                        num_positions.append((w["x0"], abs(n), w["text"]))

                # Sort by X (left → right = earlier → later columns)
                num_positions.sort(key=lambda x: x[0])
                nums = [n for _, n, _ in num_positions]

                print(f"  {sec_name}: found {len(nums)} numbers: {[round(n,2) for n in nums]}")

                if sec_name not in holdings:
                    holdings[sec_name] = {"name": sec_name}

                h = holdings[sec_name]

                # Quantity: a small integer (typically 1–5000)
                for _, n, raw in num_positions:
                    if n == int(n) and 1 <= n <= 9999 and "." not in raw:
                        h["quantity"] = int(n)
                        break

                # The last 4 numbers in the row are typically:
                # opening_ccy, opening_azn, closing_ccy, closing_azn
                if len(nums) >= 4:
                    h["opening_ccy"]  = nums[-4]
                    h["opening_azn"]  = nums[-3]
                    h["closing_ccy"]  = nums[-2]
                    h["closing_azn"]  = nums[-1]
                elif len(nums) >= 2:
                    h["opening_ccy"]  = nums[-2]
                    h["closing_ccy"]  = nums[-1]
                elif len(nums) == 1:
                    h["opening_ccy"]  = nums[0]

    return list(holdings.values())


# ── Excel update ─────────────────────────────────────────────────────────────

def find_cell(ws, search_text, max_row=120):
    """Find first cell whose value contains search_text (case-insensitive)."""
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if cell.value and search_text.lower() in str(cell.value).lower():
                return cell
    return None


def update_excel(template_path, holdings, output_path):
    wb = openpyxl.load_workbook(template_path)

    # Pick sheet: prefer one named FINAL
    ws = wb.active
    for name in wb.sheetnames:
        if "FINAL" in name.upper():
            ws = wb[name]
            break
    print(f"  Sheet: {ws.title}")

    # Locate the Holdings header row
    name_header = find_cell(ws, "Name")
    if not name_header:
        print("  ERROR: Could not find 'Name' header in Excel sheet.")
        print("  Sheets available:", wb.sheetnames)
        return

    name_col    = name_header.column
    header_row  = name_header.row

    # Map column headers → column numbers
    col_map = {}
    for cell in ws[header_row]:
        if not cell.value:
            continue
        v = str(cell.value).lower().strip()
        if "name"             in v:  col_map["name"]            = cell.column
        elif "quantity"       in v:  col_map["quantity"]        = cell.column
        elif "qty"            in v:  col_map["quantity"]        = cell.column
        elif "executed"       in v:  col_map["executed_price"]  = cell.column
        elif "trading value"  in v:  col_map["trading_value"]   = cell.column
        elif "current price"  in v:  col_map["current_price"]   = cell.column
        elif "current value"  in v:  col_map["current_value"]   = cell.column

    print(f"  Holdings header at row {header_row}, columns: {col_map}")

    updated = 0
    for holding in holdings:
        sec_name = holding["name"]

        # Find the Excel row that matches this security
        matched_row = None
        for r in range(header_row + 1, header_row + 25):
            cell_val = ws.cell(row=r, column=name_col).value
            if excel_row_matches(cell_val, sec_name):
                matched_row = r
                break

        if matched_row is None:
            print(f"  WARNING: No Excel row found for '{sec_name}'")
            continue

        print(f"  Updating row {matched_row}: {ws.cell(row=matched_row, column=name_col).value}")

        qty = holding.get("quantity", 1) or 1

        if "quantity" in col_map and "quantity" in holding:
            ws.cell(row=matched_row, column=col_map["quantity"]).value = holding["quantity"]

        if "executed_price" in col_map and "opening_ccy" in holding:
            ws.cell(row=matched_row, column=col_map["executed_price"]).value = round(
                holding["opening_ccy"] / qty, 2
            )

        if "trading_value" in col_map and "opening_ccy" in holding:
            ws.cell(row=matched_row, column=col_map["trading_value"]).value = holding["opening_ccy"]

        if "current_price" in col_map and "closing_ccy" in holding:
            ws.cell(row=matched_row, column=col_map["current_price"]).value = round(
                holding["closing_ccy"] / qty, 2
            )

        if "current_value" in col_map and "closing_ccy" in holding:
            ws.cell(row=matched_row, column=col_map["current_value"]).value = holding["closing_ccy"]

        updated += 1

    wb.save(output_path)
    print(f"\n  Updated {updated}/{len(holdings)} holdings")
    print(f"  Saved: {output_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Usage: python pdf_to_excel.py <input.pdf> <template.xlsx> [output.xlsx]")
        sys.exit(1)

    pdf_path      = sys.argv[1]
    template_path = sys.argv[2]

    if len(sys.argv) > 3:
        output_path = sys.argv[3]
    else:
        pdf_dir  = os.path.dirname(os.path.abspath(pdf_path))
        pdf_stem = os.path.splitext(os.path.basename(pdf_path))[0]
        output_path = os.path.join(pdf_dir, f"{pdf_stem}_portfolio.xlsx")

    if not os.path.exists(pdf_path):
        print(f"ERROR: PDF not found: {pdf_path}")
        sys.exit(1)
    if not os.path.exists(template_path):
        print(f"ERROR: Template not found: {template_path}")
        sys.exit(1)

    # DEBUG: print raw PDF text to see what's being extracted
    print("=== DEBUG: Raw PDF rows ===")
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            words_sorted = sorted(words, key=lambda w: (round(w["top"] / 3) * 3, w["x0"]))
            rows = group_words_into_rows(words_sorted, y_tolerance=5)
            print(f"\n--- Page {page_num + 1}: {len(rows)} rows, {len(words)} words ---")
            for i, row in enumerate(rows[:60]):  # print first 60 rows
                row_text = " ".join(w["text"] for w in row)
                print(f"  row {i:02d}: {row_text[:120]}")
    print("\n=== END DEBUG ===\n")

    print(f"Extracting from: {pdf_path}\n")
    holdings = extract_pdf_data(pdf_path)

    print(f"\nExtracted {len(holdings)} holdings:")
    for h in holdings:
        print(f"  {h}")

    print(f"\nUpdating: {template_path}")
    update_excel(template_path, holdings, output_path)
    print("\nDone!")


if __name__ == "__main__":
    main()
